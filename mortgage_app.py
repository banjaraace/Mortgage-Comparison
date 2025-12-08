import streamlit as st
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.utils import get_column_letter

st.title("Mortgage Scenario Comparison Tool")

# Scenario container
if "scenarios" not in st.session_state:
    st.session_state.scenarios = []

st.header("Add a Scenario")

with st.form("add_scenario"):
    name = st.text_input("Scenario Name")

    property_price = st.number_input("Property Price ($)", min_value=0.0, value=400000.0)
    down_payment_percent = st.number_input("Down Payment (%)", min_value=0.0, value=20.0)
    down_payment_amount = property_price * (down_payment_percent / 100)
    loan_amount = property_price - down_payment_amount

    st.write(f"**Calculated Down Payment:** ${down_payment_amount:,.2f}")
    st.write(f"**Calculated Loan Amount:** ${loan_amount:,.2f}")

    annual_rate = st.number_input("Annual Interest Rate (%)", min_value=0.0, value=5.0)
    years = st.number_input("Loan Term (Years)", min_value=1, value=30)

    extra_payment = st.number_input("Extra Monthly Payment ($)", min_value=0.0, value=0.0)
    extra_payment_months = st.number_input("Number of Months to Apply Extra Payment", min_value=0, value=0)

    submit = st.form_submit_button("Add Scenario")

if submit and name:
    st.session_state.scenarios.append({
        "name": name,
        "property_price": property_price,
        "down_payment_percent": down_payment_percent,
        "down_payment_amount": down_payment_amount,
        "loan_amount": loan_amount,
        "annual_rate": annual_rate,
        "years": years,
        "extra_payment": extra_payment,
        "extra_payment_months": extra_payment_months
    })
    st.success(f"Scenario '{name}' added!")

# ---------------------- Mortgage Calculation ----------------------------

def calculate_schedule(loan_amount, annual_rate, years, extra_payment, extra_payment_months):
    monthly_rate = annual_rate / 100 / 12
    total_months = int(years * 12)

    if monthly_rate > 0:
        monthly_payment = loan_amount * (monthly_rate * (1 + monthly_rate) ** total_months) / ((1 + monthly_rate) ** total_months - 1)
    else:
        monthly_payment = loan_amount / total_months

    balance = loan_amount
    amortization = []

    for month in range(1, total_months + 1):

        if month <= extra_payment_months:
            applied_extra = extra_payment
        else:
            applied_extra = 0.0

        interest = balance * monthly_rate
        principal = monthly_payment - interest
        principal += applied_extra

        if principal > balance:
            principal = balance
            monthly_payment_adjusted = interest + principal
        else:
            monthly_payment_adjusted = monthly_payment + applied_extra

        balance -= principal

        amortization.append({
            "Month": month,
            "Interest": round(interest, 2),
            "Principal": round(principal - applied_extra, 2),
            "Extra Payment": round(applied_extra, 2),
            "Total Payment": round(monthly_payment_adjusted, 2),
            "Remaining Balance": round(balance, 2)
        })

        if balance <= 0:
            break

    df = pd.DataFrame(amortization)

    totals = {
        "total_interest": df["Interest"].sum(),
        "total_principal": df["Principal"].sum(),
        "total_extra": df["Extra Payment"].sum(),
        "total_paid": df["Total Payment"].sum(),
        "months": len(df)
    }

    return df, totals, monthly_payment


# ---------------------- Excel Export with Formulas ----------------------------

def create_excel(scenario, df, totals, monthly_payment):
    output = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mortgage Scenario"

    # Scenario Details
    details = {
        "Scenario Name": scenario["name"],
        "Property Price": scenario["property_price"],
        "Down Payment (%)": scenario["down_payment_percent"],
        "Down Payment Amount": scenario["down_payment_amount"],
        "Loan Amount": scenario["loan_amount"],
        "Interest Rate (%)": scenario["annual_rate"],
        "Loan Term (Years)": scenario["years"],
        "Extra Monthly Payment": scenario["extra_payment"],
        "Extra Payment Months": scenario["extra_payment_months"],
        "Base Monthly Payment": monthly_payment
    }

    row = 1
    for key, value in details.items():
        ws[f"A{row}"] = key
        ws[f"B{row}"] = value
        row += 1

    start = row + 1

    headers = ["Month", "Interest", "Principal", "Extra Payment", "Total Payment", "Remaining Balance"]
    ws.append(headers)

    for i in range(1, len(df) + 1):
        r = start + i

        month = i
        ws[f"A{r}"] = month

        if i == 1:
            balance_ref = scenario["loan_amount"]
        else:
            balance_ref = f"F{r-1}"

        rate = scenario["annual_rate"] / 100 / 12

        extra = f"=IF(A{r}<={scenario['extra_payment_months']},{scenario['extra_payment']},0)"

        ws[f"B{r}"] = f"= {balance_ref} * {rate}"
        ws[f"C{r}"] = f"= {monthly_payment} - B{r}"
        ws[f"D{r}"] = extra
        ws[f"E{r}"] = f"= {monthly_payment} + D{r}"
        ws[f"F{r}"] = f"= {balance_ref} - C{r} - D{r}"

    wb.save(output)
    output.seek(0)
    return output


# ---------------------- Display Scenarios ----------------------------

st.header("Scenarios")

if not st.session_state.scenarios:
    st.info("Add a scenario to begin.")
else:
    for index, sc in enumerate(st.session_state.scenarios):
        st.subheader(sc["name"])

        df, totals, monthly_payment = calculate_schedule(
            sc["loan_amount"],
            sc["annual_rate"],
            sc["years"],
            sc["extra_payment"],
            sc["extra_payment_months"]
        )

        colA, colB = st.columns([0.85, 0.15])

        with colB:
            if st.button(f"Delete '{sc['name']}'", key=f"delete_{index}"):
                st.session_state.scenarios.pop(index)
                st.experimental_rerun()

        with colA:
            st.write(f"**Loan Amount:** ${sc['loan_amount']:,.0f}")
            st.write(f"**Interest Rate:** {sc['annual_rate']}%")
            st.write(f"**Extra Monthly Payment:** ${sc['extra_payment']:,.2f}")
            st.write(f"**Extra Payment Months:** {sc['extra_payment_months']}")
            st.write(f"**Base Monthly Payment:** ${monthly_payment:,.2f}")

            st.dataframe(df, height=350)

            st.markdown("### Totals")
            st.write(f"**Total Interest:** ${totals['total_interest']:,.2f}")
            st.write(f"**Total Principal:** ${totals['total_principal']:,.2f}")
            st.write(f"**Total Extra Payments:** ${totals['total_extra']:,.2f}")
            st.write(f"**Total Paid:** ${totals['total_paid']:,.2f}")
            st.write(f"**Loan Payoff:** {totals['months']} months ({totals['months']/12:.2f} years)")

            excel_data = create_excel(sc, df, totals, monthly_payment)

            st.download_button(
                label="Download Scenario as Excel",
                data=excel_data,
                file_name=f"{sc['name'].replace(' ', '_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        st.markdown("---")
